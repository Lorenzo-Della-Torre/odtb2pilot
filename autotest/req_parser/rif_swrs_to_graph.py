#!/usr/bin/env python3

# VCC 2020-07-16
""" Find out if there is a good way to group requirements, given the SWRS """

import logging
import argparse
import sys
# pip install lxml
from lxml import etree as ET # pylint: disable=c-extension-no-member
import openpyxl
from neo4j import GraphDatabase

# Logging has different levels: DEBUG, INFO, WARNING, ERROR, and CRITICAL
# Set the level you want to have printout in the console.
logging.basicConfig(format=' %(message)s', stream=sys.stdout, level=logging.INFO)
logger = logging.getLogger(__name__)

OUTFOLDER = "output/"
OWN_ID = "Own UUID"
PARENT_ID = "Parent UUID"
VER_ID = "Version id"

# Constants
NS = {'rif': 'http://automotive-his.de/200706/rif'}
ID = 'ID'
LONG = 'LONG'
ATTRIB = 'ATTRIB'
SHORT = 'SHORT'
FIRSTLINE = 1

def parse_some_args():
    """Get the command line input, using the defined flags."""
    parser = argparse.ArgumentParser(description=
        'Parse Elektra rif export, and give result on preffered format.')
    parser.add_argument("--elektra", help="Elektra rif export, xml-file",
        type=str, action='store', dest='rif', required=True)
    parser.add_argument("--out", help="Name of the generated file",
        dest='out', default= OUTFOLDER + 'gen_swrs_out.xlsx')
    parser.add_argument("--out_csv", help="Name of the generated file",
        dest='outcsv', default= OUTFOLDER + 'bt_req.csv')
    parser.add_argument("--neoname", help="Username for graph db.", type=str,
        dest='neoname', default='neo4j')
    parser.add_argument("--neopw", help="Password for graph db.", type=str,
        dest='neopw', default='1234')
    ret_args = parser.parse_args()
    return ret_args

def get_spec_dict(root):
    """ Parse and get data for the special attributes """
    spectypes = root.find('rif:SPEC-TYPES', NS)

    spec_dict = {}

    for spectype in spectypes:
        tmp_dict = {}
        spec_long_name = spectype.find('rif:LONG-NAME', NS).text
        logger.debug('Spec name: %s', spec_long_name)
        spec_identifier = spectype.find('rif:IDENTIFIER', NS).text
        # Find all attribute def as well
        tmp_attrib_list = []
        for attrib_node in spectype.find('rif:SPEC-ATTRIBUTES', NS).\
                findall('rif:ATTRIBUTE-DEFINITION-SIMPLE', NS):
            datatype_string = attrib_node.find('rif:LONG-NAME', NS).text
            datatype_identifier = attrib_node.find('rif:IDENTIFIER', NS).text
            tmp_attrib_list.append((datatype_string,datatype_identifier))
        short_name = spec_long_name.split()[0]
        tmp_dict[ID] = spec_identifier
        tmp_dict[LONG] = spec_long_name
        tmp_dict[ATTRIB] = tmp_attrib_list
        tmp_dict[SHORT] = short_name
        spec_dict[short_name] = tmp_dict
    logger.debug(spec_dict)
    # for key in spec_dict:
    #     print(key)
    #     logger.info(spec_dict[key]['ATTRIB'])
    logger.debug(spec_dict['Hierarchy'])
    return spec_dict


def get_spec_obj_dict(root, type_dict):
    """ Parse and get data for the special objects """
    # pylint: disable=too-many-locals
    specobjs = root.find('rif:SPEC-OBJECTS', NS)

    # Get the IDs for the spec types we are interested in
    ## REQ and REQPROD has same attibute
    reqprod_id = type_dict['REQPROD'][ID]
    reqset_id = type_dict['REQ-SET'][ID]
    req_id = type_dict['REQ'][ID]
    folder_id = type_dict['FOLDER'][ID]

    # Temporary dicts to be ut in the big dict when done.
    tmp_reqprod_dict = {}
    tmp_reqset_dict = {}
    tmp_req_dict = {}
    tmp_folder_dict = {}

    # Attrib of interest: ID, LongName, Type, list of AttribValueSimple
    spec_obj_dict = {}

    for specobj in specobjs:
        if specobj.find('rif:TYPE/rif:SPEC-TYPE-REF', NS) is None:
            logger.info("Spec object without a type!")
            continue

        spectype_id = specobj.find('rif:TYPE/rif:SPEC-TYPE-REF', NS).text
        if spectype_id == reqprod_id:
            reqprod_dict, obj_id = get_obj_data(specobj, \
                type_dict['REQPROD'][ATTRIB])
            tmp_reqprod_dict[obj_id] = reqprod_dict

        elif spectype_id == reqset_id:
            reqset_dict, obj_id = get_obj_data(specobj, type_dict['REQ-SET'][ATTRIB])
            tmp_reqset_dict[obj_id] = reqset_dict

        elif spectype_id == req_id:
            req_dict, obj_id = get_obj_data(specobj, type_dict['REQ'][ATTRIB])
            tmp_req_dict[obj_id] = req_dict

        elif spectype_id == folder_id:
            folder_dict, obj_id = get_obj_data(specobj, type_dict['FOLDER'][ATTRIB])
            tmp_folder_dict[obj_id] = folder_dict

    spec_obj_dict['REQPROD'] = tmp_reqprod_dict
    spec_obj_dict['REQ-SET'] = tmp_reqset_dict
    spec_obj_dict['REQ'] = tmp_req_dict
    spec_obj_dict['FOLDER'] = tmp_folder_dict

    return spec_obj_dict

def get_hierarchy_dict(root, hier_info_dict):
    """ Extract how nodes are linked into a hierarchy """
    # The structure is children inside children nodes...
    # Keep track of each node's parent? Then traverse the tree?
    # Parent identifier, so each node shall have info: Parent-ID, ownID,
    #  (and any attributes of interest)
    # Own identifier is found in other nodes that are already parsed (else those most be created)
    rel_dict = {}
    hier_root = root.find('.//rif:SPEC-HIERARCHY-ROOT', NS)
    # Add check if the number is the same? Not necessary for now, assume generated data is correct.
    hier_id = hier_info_dict[ID]
    logger.debug(hier_id)
    # Keep it simple, optimize later. Find Children, and use xpath to get ID of parent.
    # Save the mapping to the spec obj (if any)
    for child_node in hier_root.findall('.//rif:CHILDREN/rif:SPEC-HIERARCHY', NS):
        #child_id = child_node.find('rif:IDENTIFIER', NS).text
        # one more dict to map identifier to spec-object?
        child_id = child_node.find('rif:OBJECT/rif:SPEC-OBJECT-REF', NS).text
        #parent_id = child_node.find('../../rif:IDENTIFIER', NS).text
        parent_id = "Maaaan"
        parent_id_cand = child_node.find('../../rif:OBJECT/rif:SPEC-OBJECT-REF', NS)
        if parent_id_cand is not None:
            parent_id = parent_id_cand.text
        #logger.info(f'Child {child_id} has parent {parent_id}')
        rel_dict[child_id] = parent_id
    #logger.warning(rel_dict)
    return rel_dict

## Helper functions for parsing data
def get_obj_data(tree_obj, attrib_tup_list):
    """ Get the data for REQ-SET, FOLDER, REQ, and REQPROD object, return dict """
    # Class, ModUsrLogName, Name, Revision, State, Variant etc
    obj_dict = {}
    # Loop over the attrib refs, and then match against given IDs.
    node_id = tree_obj.find('rif:IDENTIFIER', NS).text
    node2_list = tree_obj.findall('.//rif:ATTRIBUTE-DEFINITION-SIMPLE-REF', NS)
    for node2 in node2_list:
        for attrib, attrib_id in attrib_tup_list:
            if node2.text == attrib_id:
                node2_value = node2.find('../../rif:THE-VALUE', NS)
                obj_dict[attrib] = node2_value.text
    return obj_dict, node_id



def parse_rif_to_dicts(rif_path):
    """ Get the required data for the nodes, then return the dicts """
    spec_dict = {} # dict with the info about SPEC-TYPE nodes

    tree = ET.parse(rif_path) # pylint: disable= c-extension-no-member
    root = tree.getroot()

    # need to identify what spec type is used, in order to parse
    spec_dict = get_spec_dict(root)
    # child 2 parent dict.
    hier_dict = get_hierarchy_dict(root, spec_dict['Hierarchy'])
    specobj_dict = get_spec_obj_dict(root, spec_dict)
    #print(specobj_dict['REQ-SET'].keys())
    # Find the relations somewhere...
    # To get the chapters, use Hierachy and it should consist of folders and reqset
    # But how is REQPROD connected to REQ-SET?
    return spec_dict, specobj_dict, hier_dict

def write_xlsx_out(filepath, sp_dict, spobj_dict, child_to_parent_dict):
    """ Put data in a csv file which can be opened in Excel """
    # pylint: disable=too-many-locals
    type_list = ["REQPROD", "REQ", "REQ-SET", "FOLDER"]
    # Need to add info about own ID, and parent's ID (if existing)

    # ATTRIB is a list, get the keys for col headings for all the different TYPEs
    col_names = create_col_names(sp_dict, type_list)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SWRS info"

    logger.info("In the writing phase...")
    for col_cnt, col_name in enumerate(col_names, 1):
        heading_cell = worksheet.cell(row=FIRSTLINE, column=col_cnt)
        heading_cell.style = 'Headline 1'
        heading_cell.value = col_name
    # Add the data...
    excel_line_cnt = FIRSTLINE
    for type_name in type_list:
        for obj_id in spobj_dict[type_name]:
            excel_line_cnt += 1
            #logger.error(spobj_dict[type_name][obj_id].keys())
            for col_cnt, col in enumerate(col_names, 1):
                write_cell = None
                if col == OWN_ID:
                    write_cell = obj_id
                elif col == PARENT_ID:
                    # Second arg to get, is the defaule value
                    write_cell = child_to_parent_dict.get(obj_id, "NOK_ID")
                elif col == VER_ID:
                    ver_id_value = spobj_dict[type_name][obj_id].get(col, "NA")
                    write_cell = f"easee:VCC_EEDM,{ver_id_value}"
                else:
                    #logger.info(spobj_dict[type_name][obj_id])
                    write_cell = spobj_dict[type_name][obj_id].get(col, "-_-")
                data_cell = worksheet.cell(row=excel_line_cnt, column=col_cnt)
                data_cell.value = write_cell

    workbook.save(filepath)
    logger.info("Written to file %s", filepath)

def create_col_names(sp_dict, type_list):
    """ Return the column names """
    col_list = ["Class", "ID", "Revision", "Variant", "Name"]
    name_set = set()
    for elektra_type in type_list:
        for tup_text, _ in sp_dict[elektra_type][ATTRIB]:
            name_set.add(tup_text)
    # If class exists, then first. Also add the own+parent id columns
    # Remove default col_list entries from the set
    for elem in col_list:
        name_set.discard(elem)
    col_list.extend(name_set)
    # add the UUIDs
    col_list.append(OWN_ID)
    col_list.append(PARENT_ID)

    return col_list

def write_minimum_bt_req(filepath, sp_dict, spobj_reqp_dict):
    """ Generate what the test report requires (REQPROD id etc) """
    BT_CSV_HEADER = "REQPROD;TEST_METHOD;ELEKTRA;LINK\n"
    logger.info("Writing data to BT REQPROD-csv...")
    with open(filepath, 'w') as csv_file:
        csv_file.write(BT_CSV_HEADER)
        for obj_id in spobj_reqp_dict:
            # REQPROD ID
            write_line = ""
            write_line += spobj_reqp_dict[obj_id].get('ID', "-_-") + ";"
            # Test method from categorization (not available in this scope)
            write_line += "-;"
            # Elektra test method is not currently being extracted from the arxml
            write_line += "-;"
            # Elektra link
            ver_id_value = spobj_reqp_dict[obj_id].get("Version id", "-_-")
            write_line += f"easee:VCC_EEDM,{ver_id_value}\n"
            csv_file.write(write_line)
    logger.info("Data has been written to %s", filepath)

###### Neo4j stuff ######
def neo_fix(node_dict, c2p_dict, neousr, neopw):
    """ Interact with the graph db """
    uri = "bolt://localhost:7687"
    driver = GraphDatabase.driver(uri, auth=(neousr, neopw))

    def create_elektranode(neo_tx, el_type, attr):
        neo_tx.run(f"MERGE (a:{wash_char(el_type)} {{ {attr} }})")

    def create_elektrarel(neo_tx, node, rel, other):
        neo_tx.run\
        (f"MATCH (a {{UUID: '{node}'}})\nMATCH (b {{UUID: '{other}'}})\nMERGE (a)-[r:{rel}]->(b)")

    for el_type in node_dict:
        logger.info('Neo4j working with %s', el_type)
        for obj in node_dict[el_type]:
            with driver.session() as session:
                attrib_str = get_attib_str(node_dict[el_type][obj], obj)
                session.write_transaction(create_elektranode, el_type, attrib_str)
    for child in c2p_dict:
        with driver.session() as session:
            session.write_transaction(create_elektrarel, child, 'HAS_PARENT', c2p_dict[child])

    driver.close()

def wash_char(intext):
    """ Neo4j doesn't like - in the Node type, make sure to remove """
    return intext.replace('-', '_')

def get_attib_str(obj_dict, obj_id):
    """ Dict with attributes, return list of attributes compatible with cypher """
    attr_list = [f"UUID: '{obj_id}'"]
    for key in obj_dict:
        attr_list.append(f"{key.replace(' ', '_')}: '{obj_dict[key]}'")
    return ",".join(attr_list)

def main(margs):
    """ Main function """
    logger.debug("In main!")
    sp_dict, spobj_dict, child_to_parent_dict = parse_rif_to_dicts(margs.rif)
    # Quick hack to get the csv required by test report generator
    write_minimum_bt_req(margs.outcsv, sp_dict['REQPROD']['ATTRIB'], spobj_dict["REQPROD"])

    # Map the data per node in order to generate an excel file
    # Need to set the graph adding here as well.
    # Must have FOLDER in order to create the tree structure.
    write_xlsx_out(margs.out, sp_dict, spobj_dict, child_to_parent_dict)
    neo_fix(spobj_dict, child_to_parent_dict, margs.neoname, margs.neopw)

if __name__ == "__main__":
    # Boilerplate to launch the main function with the command line arguments.
    ARGS = parse_some_args()
    main(ARGS)
